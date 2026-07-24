from __future__ import annotations

import hashlib
import hmac
import threading
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any

from cryptography.hazmat.primitives.ciphers.aead import AESGCM  # type: ignore[import-not-found]
from openpyxl import load_workbook  # type: ignore[import-untyped]

from .atomic_io import atomic_write_json
from .database_policy import (
    DatabaseTarget,
    DatabaseTargetRegistry,
    execute_sql,
    open_database_connection,
)
from .environment_guard import EnvironmentKind, ExecutionTargetPolicy
from .errors import ErrorCode, RuntimePolicyError
from .excel import AUDIT_SHEETS, write_audit_workbook
from .http_runner import HttpRequest, send_http_request
from .parsers import parse_interface_file
from .paths import WritePolicy, initialize_runtime_layout
from .state_store import load_project_state, resume_state, save_project_state

REQUIRED_SELF_CHECKS = (
    "runner_start",
    "http_local",
    "openapi_parse",
    "postman_parse",
    "excel_four_sheet",
    "state_resume",
    "sqlite",
    "database_driver",
    "crypto",
    "audit_directory_write",
    "business_repository_write_rejected",
)


def run_self_check(runtime_root: Path, fixtures_root: Path) -> dict[str, Any]:
    checks: dict[str, dict[str, str]] = {}

    def check(name: str, operation: Any) -> None:
        try:
            operation()
        except Exception as exc:
            checks[name] = {"status": "failed", "detail": f"{type(exc).__name__}: {exc}"}
        else:
            checks[name] = {"status": "passed"}

    layout = initialize_runtime_layout(runtime_root)
    write_policy = WritePolicy(
        runtime_root,
        (runtime_root.parent / "business-repository-self-check",),
    )
    self_check_root = layout["projects"] / "_self_check"
    self_check_root.mkdir(parents=True, exist_ok=True)

    check("runner_start", lambda: __import__("multi_source_test_audit.runner"))
    check("http_local", _check_local_http)
    check(
        "openapi_parse",
        lambda: _require_records(fixtures_root / "openapi.json"),
    )
    check(
        "postman_parse",
        lambda: _require_records(fixtures_root / "postman.json"),
    )
    check("excel_four_sheet", lambda: _check_excel(self_check_root, write_policy))
    check("state_resume", lambda: _check_state(runtime_root, write_policy))
    check("sqlite", _check_sqlite)
    check("database_driver", lambda: __import__("sqlite3"))
    check("crypto", _check_crypto)
    check("audit_directory_write", lambda: _check_audit_write(self_check_root, write_policy))
    check(
        "business_repository_write_rejected",
        lambda: _check_business_repository_rejection(write_policy),
    )
    passed = all(item["status"] == "passed" for item in checks.values())
    status = "ready" if passed else "incomplete"
    return {"status": status, "offline": True, "checks": checks}


def _require_records(path: Path) -> None:
    if not parse_interface_file(path):
        raise ValueError(f"No interface records parsed from {path.name}.")


def _check_local_http() -> None:
    class Handler(BaseHTTPRequestHandler):
        def do_GET(self) -> None:
            body = b'{"self_check":true}'
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        def log_message(self, format: str, *args: object) -> None:
            return

    server = ThreadingHTTPServer(("127.0.0.1", 0), Handler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        port = server.server_address[1]
        response = send_http_request(
            HttpRequest("GET", f"http://127.0.0.1:{port}/self-check"),
            ExecutionTargetPolicy(
                environment=EnvironmentKind.TEST,
                allowed_origins=frozenset({f"http://127.0.0.1:{port}"}),
                production_origins=frozenset(),
            ),
        )
        if response.status_code != 200 or response.body != {"self_check": True}:
            raise ValueError("Local HTTP self-check returned an unexpected response.")
    finally:
        server.shutdown()
        server.server_close()
        thread.join(timeout=5)


def _check_excel(root: Path, policy: WritePolicy) -> None:
    target = root / "self-check.xlsx"
    rows = {sheet: [{"status": "passed"}] for sheet in AUDIT_SHEETS}
    write_audit_workbook(target, rows, policy=policy)
    workbook = load_workbook(target, read_only=True)
    try:
        if workbook.sheetnames != list(AUDIT_SHEETS):
            raise ValueError("Excel sheet contract mismatch.")
    finally:
        workbook.close()


def _check_state(runtime_root: Path, policy: WritePolicy) -> None:
    context = {"commit": "self-check", "environment": "test", "account_alias": "self-check"}
    state = {
        "material_index": [],
        "profile": {},
        "context": context,
        "chains": [],
        "clues": [],
        "approvals": [],
        "execution": {"phase": "self-check", "status": "running"},
        "evidence_index": [],
    }
    save_project_state("_self_check", state, policy)
    loaded = load_project_state("_self_check", runtime_root)
    if resume_state(loaded, context)["requires_revalidation"]:
        raise ValueError("Unchanged self-check state was marked stale.")


def _check_sqlite() -> None:
    bound = open_database_connection(
        "self-check",
        DatabaseTargetRegistry(
            (DatabaseTarget("self-check", EnvironmentKind.TEST, ":memory:"),)
        ),
    )
    try:
        if execute_sql(bound, "SELECT 1") != [(1,)]:
            raise ValueError("SQLite returned an unexpected result.")
    finally:
        bound.connection.close()


def _check_crypto() -> None:
    key = AESGCM.generate_key(bit_length=128)
    cipher = AESGCM(key)
    nonce = b"self-check12"
    plaintext = b"multi-source-test-audit"
    ciphertext = cipher.encrypt(nonce, plaintext, b"self-check")
    if cipher.decrypt(nonce, ciphertext, b"self-check") != plaintext:
        raise ValueError("AES round trip failed.")
    digest = hashlib.sha256(plaintext).digest()
    signature = hmac.digest(key, digest, "sha256")
    if not hmac.compare_digest(signature, hmac.digest(key, digest, "sha256")):
        raise ValueError("Hash/signature round trip failed.")


def _check_audit_write(root: Path, policy: WritePolicy) -> None:
    target = root / "write-check.json"
    atomic_write_json(target, {"status": "passed"}, policy=policy)
    if not target.is_file():
        raise ValueError("Audit directory write did not create the expected file.")


def _check_business_repository_rejection(policy: WritePolicy) -> None:
    repository = policy.business_repositories[0]
    target = repository / "forbidden.json"
    try:
        atomic_write_json(target, {"status": "forbidden"}, policy=policy)
    except RuntimePolicyError as exc:
        if exc.code is ErrorCode.BUSINESS_REPOSITORY_WRITE:
            return
        raise
    finally:
        if target.exists():
            raise ValueError("Business repository write created an artifact.")
    raise ValueError("Business repository write was not rejected.")
