from __future__ import annotations

import os
import uuid
from contextlib import suppress
from pathlib import Path
from typing import Any

from .errors import ErrorCode, RuntimePolicyError
from .paths import WritePolicy, locked_write_directory, prepare_write_target
from .redaction import redact_cell

AUDIT_SHEETS = ("审计总览", "多源关联", "审计计划", "执行结果")


def write_audit_workbook(
    target: Path,
    rows_by_sheet: dict[str, list[dict[str, Any]]],
    *,
    policy: WritePolicy,
) -> None:
    if tuple(rows_by_sheet) != AUDIT_SHEETS:
        raise RuntimePolicyError(
            ErrorCode.EXCEL_SHEET_CONTRACT,
            "Audit workbook must contain exactly the four required sheets in order.",
            details={"expected": list(AUDIT_SHEETS), "actual": list(rows_by_sheet)},
        )
    try:
        from openpyxl import Workbook  # type: ignore[import-untyped]
    except ImportError as exc:
        raise RuntimePolicyError(
            ErrorCode.EXCEL_DEPENDENCY_MISSING,
            "The bundled Excel dependency is unavailable; run the repair workflow.",
        ) from exc
    guard = prepare_write_target(target, policy)
    safe_target = guard.target
    temporary_name = f".{safe_target.name}.{uuid.uuid4().hex}.tmp"
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in AUDIT_SHEETS:
        worksheet = workbook.create_sheet(sheet_name)
        rows = rows_by_sheet[sheet_name]
        headers = list(dict.fromkeys(key for row in rows for key in row))
        if headers:
            worksheet.append(headers)
            for row in rows:
                worksheet.append([redact_cell(header, row.get(header)) for header in headers])
    _save_workbook(workbook, guard, policy, temporary_name)


def write_stage_a_workbook_v2(
    target: Path,
    rows_by_sheet: dict[str, list[dict[str, Any]]],
    *,
    policy: WritePolicy,
    project_name: str,
    selected_chain: str,
) -> None:
    """Write the final Stage A workbook with deterministic professional styling."""
    if tuple(rows_by_sheet) != AUDIT_SHEETS:
        raise RuntimePolicyError(
            ErrorCode.EXCEL_SHEET_CONTRACT,
            "Audit workbook must contain exactly the four required sheets in order.",
            details={"expected": list(AUDIT_SHEETS), "actual": list(rows_by_sheet)},
        )
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (  # type: ignore[import-untyped]
            Alignment,
            Border,
            Font,
            PatternFill,
            Side,
        )
        from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
    except ImportError as exc:
        raise RuntimePolicyError(
            ErrorCode.EXCEL_DEPENDENCY_MISSING,
            "The bundled Excel dependency is unavailable; run the repair workflow.",
        ) from exc

    guard = prepare_write_target(target, policy)
    safe_target = guard.target
    temporary_name = f".{safe_target.name}.{uuid.uuid4().hex}.tmp"
    workbook = Workbook()
    workbook.remove(workbook.active)
    title_fill = PatternFill("solid", fgColor="17365D")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    status_fill = PatternFill("solid", fgColor="D9EAF7")
    pending_fill = PatternFill("solid", fgColor="FFF2CC")
    risk_fill = PatternFill("solid", fgColor="FCE4D6")
    border = Border(bottom=Side(style="thin", color="A6A6A6"))

    for sheet_name in AUDIT_SHEETS:
        worksheet = workbook.create_sheet(sheet_name)
        rows = rows_by_sheet[sheet_name]
        headers = list(dict.fromkeys(key for row in rows for key in row))
        last_column = max(1, len(headers))
        last_letter = get_column_letter(last_column)
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
        title = worksheet.cell(1, 1, f"{project_name} · 阶段 A 审计报告 · {sheet_name}")
        title.fill = title_fill
        title.font = Font(color="FFFFFF", bold=True, size=16)
        title.alignment = Alignment(horizontal="left", vertical="center")
        worksheet.row_dimensions[1].height = 30
        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
        subtitle = worksheet.cell(
            2,
            1,
            f"阶段状态：阶段 A 已完成（未执行动态验证）    已选业务链：{selected_chain}",  # noqa: RUF001
        )
        subtitle.fill = status_fill
        subtitle.font = Font(color="17365D", bold=True, size=11)
        subtitle.alignment = Alignment(vertical="center", wrap_text=True)
        worksheet.row_dimensions[2].height = 26

        if headers:
            for column, header in enumerate(headers, start=1):
                cell = worksheet.cell(4, column, header)
                cell.fill = header_fill
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
            worksheet.row_dimensions[4].height = 30
            for row_number, row in enumerate(rows, start=5):
                for column, header in enumerate(headers, start=1):
                    cell = worksheet.cell(
                        row_number,
                        column,
                        redact_cell(header, row.get(header)),
                    )
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.border = border
                    text = str(cell.value or "")
                    if "规则待确认" in text or "规则未知" in text:
                        cell.fill = pending_fill
                    elif "风险" in header and text not in {"", "无"}:
                        cell.fill = risk_fill
                    elif text == "未执行":
                        cell.fill = status_fill
                worksheet.row_dimensions[row_number].height = 42
            worksheet.auto_filter.ref = f"A4:{last_letter}{len(rows) + 4}"
            worksheet.freeze_panes = "A5"
            for column, header in enumerate(headers, start=1):
                values = [str(row.get(header, "") or "") for row in rows]
                width = max([len(str(header)) * 2, *(min(len(value), 40) for value in values)])
                worksheet.column_dimensions[get_column_letter(column)].width = min(
                    40, max(12, width + 2)
                )
        worksheet.sheet_view.showGridLines = False
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.orientation = "landscape"
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.print_title_rows = "1:4"
        worksheet.print_area = f"A1:{last_letter}{max(4, len(rows) + 4)}"

    _save_workbook(workbook, guard, policy, temporary_name)


def _save_workbook(
    workbook: Any,
    guard: Any,
    policy: WritePolicy,
    temporary_name: str,
) -> None:
    safe_target = guard.target
    with locked_write_directory(guard, policy) as directory:
        try:
            with directory.open_binary_exclusive(temporary_name) as stream:
                workbook.save(stream)
                stream.flush()
                os.fsync(stream.fileno())
            directory.replace(temporary_name, safe_target.name)
        except Exception:
            with suppress(OSError):
                directory.unlink(temporary_name)
            raise
