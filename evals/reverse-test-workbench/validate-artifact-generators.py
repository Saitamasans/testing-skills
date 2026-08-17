from pathlib import Path
import json
import subprocess
import sys
import tempfile

from docx import Document
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
SKILL = ROOT / "skills/reverse-test-workbench"
SCHEMA = SKILL / "assets/run-data.schema.json"
CONTRACT = SKILL / "references/run-data-contract.md"
VALIDATOR = SKILL / "scripts/validate_run_data.py"
BUILDER = SKILL / "scripts/build_artifacts.py"
INITIALIZER = SKILL / "scripts/init_run_data.py"
CONSISTENCY = SKILL / "scripts/check_artifact_consistency.py"
FIXTURE = ROOT / "evals/reverse-test-workbench/fixtures/minimal-run-data.json"

EXPECTED_SHEETS = [
    "01_批次与运行状态",
    "02_功能菜单清单",
    "03_页面字段按钮清单",
    "04_执行路径清单",
    "05_测试用例清单",
    "06_缺陷疑似问题清单",
    "07_风险与待确认清单",
    "08_测试数据台账",
    "09_截图证据索引",
    "10_认知资产清单",
]


for path in (SCHEMA, CONTRACT, VALIDATOR, BUILDER, INITIALIZER, CONSISTENCY, FIXTURE):
    if not path.exists():
        raise AssertionError(f"required artifact generator resource missing: {path}")

for script in (VALIDATOR, BUILDER):
    script_text = script.read_text(encoding="utf-8")
    for forbidden in ("C:\\\\", "/Users/", "Path.home()", "USERPROFILE"):
        if forbidden in script_text:
            raise AssertionError(f"platform-specific path found in {script}: {forbidden}")

schema = json.loads(SCHEMA.read_text(encoding="utf-8"))
if schema.get("$id") != "reverse-test-workbench/run-data.schema.json":
    raise AssertionError("unexpected run-data schema id")

fixture_before = FIXTURE.read_bytes()
subprocess.run([sys.executable, str(VALIDATOR), str(FIXTURE)], check=True)

with tempfile.TemporaryDirectory() as temp_dir:
    initialized = Path(temp_dir) / "evidence/run-data.json"
    subprocess.run(
        [
            sys.executable,
            str(INITIALIZER),
            "--output",
            str(initialized),
            "--run-id",
            "init-run-001",
            "--system-name",
            "初始化样例",
            "--target-url",
            "https://example.test/admin",
            "--skill-version",
            "0.1.0",
            "--executor-version",
            "0.0.79",
            "--started-at",
            "2026-08-15T11:30:00+08:00",
            "--budget-minutes",
            "8",
        ],
        check=True,
    )
    subprocess.run([sys.executable, str(VALIDATOR), str(initialized)], check=True)
    initialized_data = json.loads(initialized.read_text(encoding="utf-8"))
    if initialized_data["run"]["executor"]["protocol"] != "official-playwright-mcp":
        raise AssertionError("initializer did not create canonical executor fields")
    budget = initialized_data["run"]["budget_control"]
    if budget["closeout_at"] >= budget["deadline_at"]:
        raise AssertionError("initializer did not reserve time before the hard deadline")

with tempfile.TemporaryDirectory() as temp_dir:
    output_dir = Path(temp_dir)
    subprocess.run(
        [
            sys.executable,
            str(BUILDER),
            "--input",
            str(FIXTURE),
            "--output-dir",
            str(output_dir),
        ],
        check=True,
    )

    docx_path = output_dir / "过程小结.docx"
    xlsx_path = output_dir / "测试资产表.xlsx"
    state_path = output_dir / "evidence/_run-state.json"
    for path in (docx_path, xlsx_path, state_path):
        if not path.exists() or path.stat().st_size == 0:
            raise AssertionError(f"generated artifact missing or empty: {path}")

    document = Document(docx_path)
    document_text = "\n".join(p.text for p in document.paragraphs)
    for fragment in ("探索测试过程小结", "已回答问题", "未回答问题", "结论边界"):
        if fragment not in document_text:
            raise AssertionError(f"DOCX missing expected section: {fragment}")

    workbook = load_workbook(xlsx_path, data_only=False)
    if workbook.sheetnames != EXPECTED_SHEETS:
        raise AssertionError(f"unexpected sheets: {workbook.sheetnames}")
    for sheet in workbook.worksheets:
        if sheet.max_column < 2 or not sheet.cell(1, 1).value:
            raise AssertionError(f"sheet missing fixed headers: {sheet.title}")

    formula_data = json.loads(FIXTURE.read_text(encoding="utf-8"))
    formula_data["navigation"][0]["notes"] = '=HYPERLINK("https://example.invalid/leak","open")'
    formula_fixture = output_dir / "formula-run-data.json"
    formula_fixture.write_text(json.dumps(formula_data, ensure_ascii=False), encoding="utf-8")
    formula_output = output_dir / "formula-output"
    subprocess.run(
        [sys.executable, str(BUILDER), "--input", str(formula_fixture), "--output-dir", str(formula_output)],
        check=True,
    )
    formula_book = load_workbook(formula_output / "测试资产表.xlsx", data_only=False)
    formula_sheet = formula_book["02_功能菜单清单"]
    notes_column = next(
        index for index in range(1, formula_sheet.max_column + 1)
        if formula_sheet.cell(1, index).value == "备注"
    )
    formula_cell = formula_sheet.cell(2, notes_column)
    if formula_cell.data_type != "s" or formula_cell.value != formula_data["navigation"][0]["notes"]:
        raise AssertionError("XLSX report treated untrusted text as a formula")

    state = json.loads(state_path.read_text(encoding="utf-8"))
    if state.get("run_id") != "fixture-run-001":
        raise AssertionError("run-state projection does not match source run")
    if state.get("executor_protocol") != "official-playwright-mcp":
        raise AssertionError("run-state projection lost executor protocol")
    if state.get("executor_version") != "0.0.79":
        raise AssertionError("run-state projection lost executor version")
    if state.get("budget_control", {}).get("status") != "closing":
        raise AssertionError("run-state projection lost budget control state")

    subprocess.run(
        [sys.executable, str(CONSISTENCY), "--run-root", str(output_dir)],
        check=True,
    )
    published_run_data = output_dir / "evidence/run-data.json"
    published = json.loads(published_run_data.read_text(encoding="utf-8"))
    published["summary"]["one_line_conclusion"] = "changed after artifact publication"
    published_run_data.write_text(
        json.dumps(published, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    stale = subprocess.run(
        [sys.executable, str(CONSISTENCY), "--run-root", str(output_dir)],
        capture_output=True,
        text=True,
    )
    if stale.returncode == 0 or "stale" not in (stale.stdout + stale.stderr):
        raise AssertionError("consistency checker accepted stale derived artifacts")

    repaired = subprocess.run(
        [
            sys.executable,
            str(BUILDER),
            "--input",
            str(FIXTURE),
            "--output-dir",
            str(output_dir),
        ],
        check=True,
        capture_output=True,
        text=True,
    )
    if json.loads(repaired.stdout).get("status") != "generated":
        raise AssertionError("builder did not repair a tampered canonical run-data file")
    repeat = subprocess.run(
        [
            sys.executable,
            str(BUILDER),
            "--input",
            str(FIXTURE),
            "--output-dir",
            str(output_dir),
        ],
        check=True,
        capture_output=True,
        text=True,
    )
    if json.loads(repeat.stdout).get("status") != "unchanged":
        raise AssertionError("unchanged verified artifacts triggered a redundant full rebuild")
    if not docx_path.exists() or not xlsx_path.exists():
        raise AssertionError("repeat generation removed derived artifacts")

    xlsx_path.write_bytes(b"")
    corrupted = subprocess.run(
        [sys.executable, str(CONSISTENCY), "--run-root", str(output_dir)],
        capture_output=True,
        text=True,
    )
    if corrupted.returncode == 0 or "corrupted" not in (corrupted.stdout + corrupted.stderr):
        raise AssertionError("consistency checker accepted a truncated generated artifact")

with tempfile.TemporaryDirectory() as temp_dir:
    output_dir = Path(temp_dir)
    subprocess.run(
        [
            sys.executable,
            str(BUILDER),
            "--input",
            str(FIXTURE),
            "--output-dir",
            str(output_dir),
            "--state-only",
        ],
        check=True,
    )
    if (output_dir / "过程小结.docx").exists() or (output_dir / "测试资产表.xlsx").exists():
        raise AssertionError("state-only mode generated heavy artifacts")
    for path in (
        output_dir / "evidence/run-data.json",
        output_dir / "evidence/_run-state.json",
    ):
        if not path.exists() or path.stat().st_size == 0:
            raise AssertionError(f"state-only output missing: {path}")

if FIXTURE.read_bytes() != fixture_before:
    raise AssertionError("generator modified its input JSON")

with tempfile.TemporaryDirectory() as temp_dir:
    bad_path = Path(temp_dir) / "bad-key-findings.json"
    bad = json.loads(FIXTURE.read_text(encoding="utf-8"))
    bad["summary"]["key_findings"] = ["plain strings are not report findings"]
    bad_path.write_text(json.dumps(bad, ensure_ascii=False), encoding="utf-8")
    result = subprocess.run(
        [sys.executable, str(VALIDATOR), str(bad_path)],
        capture_output=True,
        text=True,
    )
    if result.returncode == 0 or "key_findings" not in (result.stdout + result.stderr):
        raise AssertionError("validator accepted non-object key findings")

with tempfile.TemporaryDirectory() as temp_dir:
    bad_path = Path(temp_dir) / "missing-budget-control.json"
    bad = json.loads(FIXTURE.read_text(encoding="utf-8"))
    del bad["run"]["budget_control"]
    bad_path.write_text(json.dumps(bad, ensure_ascii=False), encoding="utf-8")
    result = subprocess.run(
        [sys.executable, str(VALIDATOR), str(bad_path)],
        capture_output=True,
        text=True,
    )
    if result.returncode == 0 or "budget_control" not in (result.stdout + result.stderr):
        raise AssertionError("validator accepted missing budget control")

with tempfile.TemporaryDirectory() as temp_dir:
    bad_path = Path(temp_dir) / "invalid-budget-window.json"
    bad = json.loads(FIXTURE.read_text(encoding="utf-8"))
    bad["run"]["budget_control"]["closeout_at"] = "2026-08-15T11:50:00+08:00"
    bad_path.write_text(json.dumps(bad, ensure_ascii=False), encoding="utf-8")
    result = subprocess.run(
        [sys.executable, str(VALIDATOR), str(bad_path)],
        capture_output=True,
        text=True,
    )
    if result.returncode == 0 or "closeout_at" not in (result.stdout + result.stderr):
        raise AssertionError("validator accepted closeout_at after deadline_at")

with tempfile.TemporaryDirectory() as temp_dir:
    bad_path = Path(temp_dir) / "noncanonical-executor.json"
    bad = json.loads(FIXTURE.read_text(encoding="utf-8"))
    executor = bad["run"]["executor"]
    executor["executor_protocol"] = executor.pop("protocol")
    executor["executor_version"] = executor.pop("version")
    bad_path.write_text(json.dumps(bad, ensure_ascii=False), encoding="utf-8")
    result = subprocess.run(
        [sys.executable, str(VALIDATOR), str(bad_path)],
        capture_output=True,
        text=True,
    )
    if result.returncode == 0 or "executor" not in (result.stdout + result.stderr):
        raise AssertionError("validator accepted noncanonical executor fields")

with tempfile.TemporaryDirectory() as temp_dir:
    bad_path = Path(temp_dir) / "missing-timing-phase.json"
    bad = json.loads(FIXTURE.read_text(encoding="utf-8"))
    del bad["run"]["timing"]["executor_gate"]
    bad_path.write_text(json.dumps(bad, ensure_ascii=False), encoding="utf-8")
    result = subprocess.run(
        [sys.executable, str(VALIDATOR), str(bad_path)],
        capture_output=True,
        text=True,
    )
    if result.returncode == 0 or "timing" not in (result.stdout + result.stderr):
        raise AssertionError("validator accepted incomplete timing phases")

with tempfile.TemporaryDirectory() as temp_dir:
    bad_path = Path(temp_dir) / "bad.json"
    bad = json.loads(FIXTURE.read_text(encoding="utf-8"))
    bad["run"]["password"] = "must-not-leak"
    bad_path.write_text(json.dumps(bad, ensure_ascii=False), encoding="utf-8")
    result = subprocess.run(
        [sys.executable, str(VALIDATOR), str(bad_path)],
        capture_output=True,
        text=True,
    )
    if result.returncode == 0 or "sensitive key" not in (result.stdout + result.stderr):
        raise AssertionError("validator accepted sensitive data")

with tempfile.TemporaryDirectory() as temp_dir:
    bad_path = Path(temp_dir) / "bad-sensitive-value.json"
    bad = json.loads(FIXTURE.read_text(encoding="utf-8"))
    bad["summary"]["one_line_conclusion"] = "Authorization: Bearer rtw-review-canary-secret-value"
    bad_path.write_text(json.dumps(bad, ensure_ascii=False), encoding="utf-8")
    result = subprocess.run(
        [sys.executable, str(VALIDATOR), str(bad_path)],
        capture_output=True,
        text=True,
    )
    if result.returncode == 0 or "sensitive value" not in (result.stdout + result.stderr):
        raise AssertionError("validator accepted a credential-shaped value in a normal field")

with tempfile.TemporaryDirectory() as temp_dir:
    bad_path = Path(temp_dir) / "bad-path.json"
    bad = json.loads(FIXTURE.read_text(encoding="utf-8"))
    bad["evidence"][0]["file_path"] = "C:/private/evidence.png"
    bad_path.write_text(json.dumps(bad, ensure_ascii=False), encoding="utf-8")
    result = subprocess.run(
        [sys.executable, str(VALIDATOR), str(bad_path)],
        capture_output=True,
        text=True,
    )
    if result.returncode == 0 or "must be relative" not in (result.stdout + result.stderr):
        raise AssertionError("validator accepted an absolute Windows evidence path")

with tempfile.TemporaryDirectory() as temp_dir:
    bad_path = Path(temp_dir) / "bad-url.json"
    bad = json.loads(FIXTURE.read_text(encoding="utf-8"))
    bad["run"]["target_url"] = "https://user:pass@example.test/admin"
    bad_path.write_text(json.dumps(bad, ensure_ascii=False), encoding="utf-8")
    result = subprocess.run(
        [sys.executable, str(VALIDATOR), str(bad_path)],
        capture_output=True,
        text=True,
    )
    if result.returncode == 0 or "embedded credentials" not in (result.stdout + result.stderr):
        raise AssertionError("validator accepted credentials embedded in target_url")

with tempfile.TemporaryDirectory() as temp_dir:
    output_dir = Path(temp_dir)
    existing_docx = output_dir / "过程小结.docx"
    existing_xlsx = output_dir / "测试资产表.xlsx"
    existing_docx.write_bytes(b"existing-docx")
    existing_xlsx.write_bytes(b"existing-xlsx")
    bad_path = output_dir / "invalid.json"
    bad_path.write_text("{}", encoding="utf-8")
    result = subprocess.run(
        [
            sys.executable,
            str(BUILDER),
            "--input",
            str(bad_path),
            "--output-dir",
            str(output_dir),
        ],
        capture_output=True,
        text=True,
    )
    if result.returncode == 0:
        raise AssertionError("builder accepted invalid run data")
    if existing_docx.read_bytes() != b"existing-docx" or existing_xlsx.read_bytes() != b"existing-xlsx":
        raise AssertionError("failed generation modified existing valid artifacts")

print("artifact generator validation passed")
