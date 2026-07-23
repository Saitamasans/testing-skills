#!/usr/bin/env python3
"""Deterministic Node-free renderer and verifier for graybox report assets."""

from __future__ import annotations

import argparse
import copy
import hashlib
import json
import os
import re
import sys
import tempfile
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError as error:  # pragma: no cover - exercised by environments without the dependency
    raise SystemExit(
        "Excel 备用生成失败：当前 Python 缺少 openpyxl；已保留报告 JSON/HTML，未生成或宣称校验通过的 Excel。"
    ) from error


REQUIREMENT_COLUMNS = [
    "用例 ID", "所属模块", "用例标题", "验证功能点", "前置条件",
    "测试步骤", "预期结果", "优先级", "执行结果（通过 / 不通过 / 未执行）", "备注",
]
LEGACY_COLUMNS = [
    "用例 ID", "所属模块", "用例标题", "验证功能点", "前置条件",
    "测试步骤", "预期结果", "优先级", "执行结果", "备注",
]
SHORT_COLUMNS = ["用例 ID", "所属模块", "用例标题", "前置条件", "测试步骤", "预期结果", "优先级", "执行结果"]
WORKBENCH_COLUMNS = [*LEGACY_COLUMNS[:8], "实际结果", *LEGACY_COLUMNS[8:]]
STATUSES = ["未执行", "通过", "不通过", "待定"]
PRIORITIES = ["P0", "P1", "P2"]
REQUIREMENT_SKILLS = {"requirement-test-workbench", "enhanced-graybox-test-case-generation"}


def primary_skill(data):
    invocation = data.get("skill_invocation")
    return invocation if isinstance(invocation, str) else (invocation or {}).get("primary")


def uses_requirement_contract(data):
    return data.get("case_contract") == "requirement-test-case-v1" or primary_skill(data) in REQUIREMENT_SKILLS


def normalize_report(data):
    if not uses_requirement_contract(data):
        return copy.deepcopy(data)
    normalized = copy.deepcopy(data)
    next_id = 1
    supported = [SHORT_COLUMNS, LEGACY_COLUMNS, WORKBENCH_COLUMNS, REQUIREMENT_COLUMNS]
    for sheet in normalized.get("sheets", []):
        if sheet.get("kind") != "test_cases" or sheet.get("columns") not in supported:
            continue
        source_columns = list(sheet["columns"])

        def value(row, column):
            try:
                return row["values"][source_columns.index(column)]
            except ValueError:
                return ""

        converted = []
        for row in sheet["rows"]:
            if row.get("divider"):
                values = [
                    value(row, "用例 ID") or "【模块分割行】", value(row, "所属模块"), value(row, "用例标题"),
                    value(row, "验证功能点") or "-", value(row, "前置条件") or "-", value(row, "测试步骤") or "-",
                    value(row, "预期结果") or "-", value(row, "优先级") or "-", "-", value(row, "备注") or "模块起始分割",
                ]
            else:
                title = str(value(row, "用例标题") or "").strip()
                values = [
                    str(next_id), value(row, "所属模块"), title, value(row, "验证功能点") or f"1. {title}",
                    value(row, "前置条件"), value(row, "测试步骤"), value(row, "预期结果"),
                    value(row, "优先级"), "未执行", value(row, "备注") or "",
                ]
                next_id += 1
            converted.append({**row, "values": values})
        sheet["columns"] = list(REQUIREMENT_COLUMNS)
        sheet["rows"] = converted
    return normalized


def validate_report(data):
    if not isinstance(data, dict):
        raise ValueError("报告必须是 JSON 对象")
    for key in ("title", "generated_at", "skill_invocation", "sheets"):
        if not data.get(key):
            raise ValueError(f"缺少字段：{key}")
    if not isinstance(data["sheets"], list) or not data["sheets"]:
        raise ValueError("sheets 不能为空")
    requirement = uses_requirement_contract(data)
    if requirement and data.get("excel_font") != "SimHei":
        raise ValueError("需求类正式用例报告必须显式使用 SimHei 字体")
    next_id = 1
    case_sheet_count = 0
    for sheet in data["sheets"]:
        columns = sheet.get("columns")
        rows = sheet.get("rows")
        if not sheet.get("name") or not isinstance(columns, list) or not isinstance(rows, list):
            raise ValueError("sheet 结构无效")
        for row in rows:
            if not isinstance(row.get("values"), list) or len(row["values"]) != len(columns):
                raise ValueError(f"{sheet['name']} 存在与表头不一致的数据")
        if sheet.get("kind") != "test_cases":
            continue
        case_sheet_count += 1
        if requirement and (sheet["name"] != "正式测试用例" or columns != REQUIREMENT_COLUMNS):
            raise ValueError("需求类正式用例必须使用正式测试用例 Sheet 和统一十列表头")
        priority_index = columns.index("优先级")
        status_index = max(columns.index("执行结果") if "执行结果" in columns else -1, columns.index(REQUIREMENT_COLUMNS[8]) if REQUIREMENT_COLUMNS[8] in columns else -1)
        for row in rows:
            if row.get("divider"):
                continue
            values = row["values"]
            if requirement and str(values[0]) != str(next_id):
                raise ValueError("需求类用例 ID 必须从 1 开始连续")
            next_id += 1
            if values[priority_index] not in PRIORITIES:
                raise ValueError(f"优先级无效：{values[priority_index]}")
            if values[status_index] not in STATUSES:
                raise ValueError(f"执行结果无效：{values[status_index]}")
            if requirement and values[status_index] != "未执行":
                raise ValueError("需求类新生成正式用例的执行结果必须默认为未执行")
    if requirement and case_sheet_count != 1:
        raise ValueError("需求类报告必须且只能包含一个正式测试用例 Sheet")


def report_id(data):
    stable = json.dumps(
        {
            "skill": data["skill_invocation"],
            "project": data.get("project") or data["title"],
            "generated_at": data["generated_at"],
            "sheets": data["sheets"],
        },
        ensure_ascii=False,
        separators=(",", ":"),
    )
    return f"testing-skills:{hashlib.sha256(stable.encode('utf-8')).hexdigest()[:24]}"


def safe_sheet_name(name, used):
    base = re.sub(r"[\\/*?:\[\]]", "_", str(name))[:31] or "Sheet"
    candidate = base
    suffix = 2
    while candidate in used:
        candidate = f"{base[:27]}_{suffix}"
        suffix += 1
    used.add(candidate)
    return candidate


def sheet_matrix(sheet):
    matrix = [list(sheet["columns"])]
    for row in sheet["rows"]:
        matrix.append(list(row["values"]))
    return matrix


def render_xlsx(data, output_path):
    workbook = Workbook()
    workbook.remove(workbook.active)
    used = set()
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    font_name = data.get("excel_font") or "Microsoft YaHei"
    widths = [15, 18, 30, 34, 34, 42, 42, 10, 24, 28]

    for declared in data["sheets"]:
        worksheet = workbook.create_sheet(safe_sheet_name(declared["name"], used))
        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = "A2"
        matrix = sheet_matrix(declared)
        for row_index, values in enumerate(matrix, start=1):
            source_row = declared["rows"][row_index - 2] if row_index > 1 else None
            for column_index, value in enumerate(values, start=1):
                cell = worksheet.cell(row=row_index, column=column_index, value=value)
                cell.font = Font(name=font_name, size=10, bold=row_index == 1 or bool(source_row and source_row.get("divider")), color="FFFFFF" if row_index == 1 else ("1F4E78" if source_row and source_row.get("divider") else "1F2937"))
                cell.fill = PatternFill("solid", fgColor="1F4E78" if row_index == 1 else ("D9EAF7" if source_row and source_row.get("divider") else "FFFFFF"))
                cell.alignment = Alignment(horizontal="center" if row_index == 1 else "left", vertical="center" if row_index == 1 or source_row and source_row.get("divider") else "top", wrap_text=True)
                cell.border = border
            worksheet.row_dimensions[row_index].height = 28 if row_index == 1 else (25 if source_row and source_row.get("divider") else 42)

        column_widths = widths if declared.get("kind") == "test_cases" and len(declared["columns"]) == 10 else [22, *([36] * (len(declared["columns"]) - 1))]
        for column_index, width in enumerate(column_widths, start=1):
            worksheet.column_dimensions[worksheet.cell(row=1, column=column_index).column_letter].width = width
        worksheet.auto_filter.ref = f"A1:{worksheet.cell(row=len(matrix), column=len(declared['columns'])).coordinate}"

        if declared.get("kind") == "test_cases":
            priority_column = declared["columns"].index("优先级") + 1
            status_column = declared["columns"].index(REQUIREMENT_COLUMNS[8]) + 1
            validation = DataValidation(type="list", formula1='"未执行,通过,不通过,待定"', allow_blank=False)
            worksheet.add_data_validation(validation)
            for row_index, source_row in enumerate(declared["rows"], start=2):
                if source_row.get("divider"):
                    continue
                validation.add(worksheet.cell(row=row_index, column=status_column))
                priority = source_row["values"][priority_column - 1]
                color = {"P0": "FCE4D6", "P1": "FFF2CC", "P2": "E2F0D9"}[priority]
                worksheet.cell(row=row_index, column=priority_column).fill = PatternFill("solid", fgColor=color)
            last_row = len(matrix)
            last_column = worksheet.cell(row=1, column=len(declared["columns"])).column_letter
            status_letter = worksheet.cell(row=1, column=status_column).column_letter
            worksheet.conditional_formatting.add(
                f"A2:{last_column}{last_row}",
                FormulaRule(formula=[f'${status_letter}2="不通过"'], fill=PatternFill("solid", fgColor="FCE8E6"), font=Font(color="7F1D1D")),
            )
            worksheet.conditional_formatting.add(
                f"A2:{last_column}{last_row}",
                FormulaRule(formula=[f'${status_letter}2="待定"'], fill=PatternFill("solid", fgColor="E5E7EB"), font=Font(color="374151")),
            )
    workbook.save(output_path)


def render_html(data, output_path):
    payload = {**data, "report_id": report_id(data)}
    payload_json = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).replace("<", "\\u003c")
    statuses_json = json.dumps(STATUSES, ensure_ascii=False, separators=(",", ":"))
    html = f'''<!doctype html><html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{data["title"]}</title><style>body{{font:14px/1.5 "Microsoft YaHei",sans-serif;margin:20px;background:#f4f7fb;color:#1f2937}}section{{background:white;margin:16px 0;overflow:auto}}table{{border-collapse:collapse;width:100%;min-width:900px}}th{{background:#1f4e78;color:white;position:sticky;top:0}}th,td{{padding:8px;border:1px solid #d9e2f3;vertical-align:top;white-space:pre-wrap}}tr.divider td{{background:#d9eaf7;color:#1f4e78;font-weight:bold}}</style></head>
<body><h1 id="title"></h1><main id="content"></main><script>const report={payload_json};const statuses={statuses_json};const esc=v=>String(v??'').replace(/[&<>\"]/g,c=>({{'&':'&amp;','<':'&lt;','>':'&gt;','\"':'&quot;'}}[c]));document.querySelector('#title').textContent=report.title;const content=document.querySelector('#content');for(const sheet of report.sheets){{const section=document.createElement('section');section.innerHTML='<h2>'+esc(sheet.name)+'</h2>';const table=document.createElement('table');table.innerHTML='<thead><tr>'+sheet.columns.map(c=>'<th>'+esc(c)+'</th>').join('')+'</tr></thead>';const body=document.createElement('tbody');for(const row of sheet.rows){{const tr=document.createElement('tr');if(row.divider)tr.className='divider';for(const value of row.values){{const td=document.createElement('td');td.innerHTML=esc(value).replace(/\\n/g,'<br>');tr.appendChild(td)}}body.appendChild(tr)}}table.appendChild(body);section.appendChild(table);content.appendChild(section)}}</script></body></html>'''
    Path(output_path).write_text(html, encoding="utf-8")


def verify_xlsx(data, workbook_path):
    workbook = load_workbook(workbook_path, data_only=False)
    expected_names = [sheet["name"] for sheet in data["sheets"]]
    if workbook.sheetnames != expected_names:
        raise ValueError(f"Excel Sheet 名称不一致：{workbook.sheetnames!r}")
    for declared in data["sheets"]:
        worksheet = workbook[declared["name"]]
        expected = sheet_matrix(declared)
        if worksheet.max_row != len(expected) or worksheet.max_column != len(declared["columns"]):
            raise ValueError(f"{declared['name']} 尺寸错误：{worksheet.max_row}x{worksheet.max_column}")
        actual = []
        for row_index in range(1, worksheet.max_row + 1):
            actual_row = []
            for column_index in range(1, worksheet.max_column + 1):
                actual_row.append(worksheet.cell(row=row_index, column=column_index).value)
            actual.append(actual_row)
        if actual != expected:
            raise ValueError(f"{declared['name']} 单元格内容与 JSON 不一致")
        if declared.get("kind") == "test_cases":
            if actual[0] != REQUIREMENT_COLUMNS or actual[0][0] != "用例 ID" or actual[0][1] != "所属模块" or actual[0][9] != "备注":
                raise ValueError("正式测试用例表头不完整")
            case_rows = [row for source, row in zip(declared["rows"], actual[1:]) if not source.get("divider")]
            if [str(row[0]) for row in case_rows] != [str(value) for value in range(1, len(case_rows) + 1)]:
                raise ValueError("正式用例 ID 不连续")
            if any(row[8] != "未执行" for row in case_rows):
                raise ValueError("正式用例执行结果不是全部未执行")


def verify_html(data, html_path):
    html = Path(html_path).read_text(encoding="utf-8")
    match = re.search(r"const report=(.*?);const statuses=", html, re.DOTALL)
    if not match:
        raise ValueError("HTML 未包含可验证的同源 JSON")
    payload = json.loads(match.group(1))
    payload.pop("report_id", None)
    if payload != data:
        raise ValueError("HTML 内容与 JSON 不一致")


def verify_assets(data, xlsx_path, html_path):
    verify_xlsx(data, xlsx_path)
    verify_html(data, html_path)


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output-dir")
    parser.add_argument("--basename")
    parser.add_argument("--verify-only", action="store_true")
    parser.add_argument("--xlsx")
    parser.add_argument("--html")
    return parser.parse_args()


def main():
    args = parse_args()
    data = normalize_report(json.loads(Path(args.input).read_text(encoding="utf-8")))
    validate_report(data)
    if args.verify_only:
        if not args.xlsx or not args.html:
            raise ValueError("--verify-only 必须同时提供 --xlsx 和 --html")
        verify_assets(data, args.xlsx, args.html)
        print(json.dumps({"verified": True, "xlsx": args.xlsx, "html": args.html}, ensure_ascii=False))
        return
    if not args.output_dir or not args.basename:
        raise ValueError("备用生成必须提供 --output-dir 和 --basename")

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = output_dir / f"{args.basename}.xlsx"
    html_path = output_dir / f"{args.basename}.html"
    with tempfile.TemporaryDirectory(prefix="graybox-render-", dir=output_dir) as temporary:
        temporary_dir = Path(temporary)
        temporary_xlsx = temporary_dir / xlsx_path.name
        temporary_html = temporary_dir / html_path.name
        render_html(data, temporary_html)
        render_xlsx(data, temporary_xlsx)
        verify_assets(data, temporary_xlsx, temporary_html)
        os.replace(temporary_html, html_path)
        os.replace(temporary_xlsx, xlsx_path)
    print(json.dumps({"xlsx": str(xlsx_path), "html": str(html_path), "verified": True, "renderer": "python-fallback"}, ensure_ascii=False))


if __name__ == "__main__":
    try:
        main()
    except Exception as error:
        print(f"Excel/HTML 生成或真实单元格校验失败：{error}", file=sys.stderr)
        raise SystemExit(1)
