import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "tooling"))
from build_skills import build_all, load_manifest, parse_frontmatter


SLUG = "enhanced-graybox-test-case-generation"
SOURCE = ROOT / "增强版-灰盒测试用例生成_Skill.md"
PACKAGE = ROOT / "skills" / SLUG
CASE_COLUMNS = [
    "用例 ID",
    "所属模块",
    "用例标题",
    "验证功能点",
    "前置条件",
    "测试步骤",
    "预期结果",
    "优先级",
    "执行结果（通过 / 不通过 / 未执行）",
    "备注",
]


def order_export_report():
    return json.loads((ROOT / "tests" / "fixtures" / "enhanced-graybox-order-export.json").read_text(encoding="utf-8"))


def expected_sheet_matrix(sheet):
    return [sheet["columns"], *[row["values"] for row in sheet["rows"]]]


def assert_workbook_matches_report(testcase, workbook_path, report):
    workbook = load_workbook(workbook_path, data_only=False)
    testcase.assertEqual([sheet["name"] for sheet in report["sheets"]], workbook.sheetnames)
    for declared in report["sheets"]:
        worksheet = workbook[declared["name"]]
        expected = expected_sheet_matrix(declared)
        testcase.assertEqual(len(expected), worksheet.max_row, declared["name"])
        testcase.assertEqual(len(declared["columns"]), worksheet.max_column, declared["name"])
        actual = [
            [worksheet.cell(row=row, column=column).value for column in range(1, worksheet.max_column + 1)]
            for row in range(1, worksheet.max_row + 1)
        ]
        testcase.assertEqual(expected, actual, declared["name"])


def assert_html_matches_report(testcase, html_path, report):
    html = Path(html_path).read_text(encoding="utf-8")
    match = re.search(r"const report=(.*?);const statuses=", html, re.DOTALL)
    testcase.assertIsNotNone(match)
    payload = json.loads(match.group(1))
    testcase.assertRegex(payload.pop("report_id"), r"^testing-skills:[0-9a-f]{24}$")
    testcase.assertEqual(report, payload)


class EnhancedGrayboxSkillContractsTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.text = SOURCE.read_text(encoding="utf-8") if SOURCE.exists() else ""

    def assert_terms(self, *terms):
        for term in terms:
            self.assertIn(term, self.text, term)

    def assert_terms_in(self, start, end, *terms):
        self.assertIn(start, self.text)
        section = self.text.split(start, 1)[1]
        if end:
            self.assertIn(end, section)
            section = section.split(end, 1)[0]
        for term in terms:
            self.assertIn(term, section, f"{start}: {term}")

    def test_01_complete_requirement_and_backend_enable_basic_graybox(self):
        self.assert_terms_in("## 双重输入分级", "## L0 与 P0 门禁", "G1", "基础灰盒增强", "正式测试用例")
        self.assert_terms("实现证据")

    def test_02_android_backend_and_contract_build_parameter_chain(self):
        self.assert_terms_in("## 需求相关全链路定位", "## 真值优先级", "客户端参数来源", "后端读取", "前后端差异", "差异进入测试点")

    def test_03_l0_with_complete_code_cannot_bypass_p0(self):
        self.assert_terms_in("## 双重输入分级", "## L0 与 P0 门禁", "L0 + G3", "不得因代码完整而生成正式测试用例")
        self.assert_terms_in("## L0 与 P0 门禁", "## 判断开发阶段", "当前实现证据", "禁止输出正式测试用例")

    def test_04_requirement_only_routes_without_code_blocker(self):
        self.assert_terms_in("## 输入准入与互斥路由", "## 双重输入分级", "只有需求文档", "requirement-test-workbench", "不扫描代码仓库", "不把“没有代码”当成需求阻塞")

    def test_05_code_only_does_not_invent_product_expectations(self):
        self.assert_terms_in("## 输入准入与互斥路由", "## 双重输入分级", "只有代码", "不编造业务预期", "代码审计", "不生成正式需求用例")

    def test_06_predevelopment_code_is_only_a_baseline(self):
        self.assert_terms_in("### A. 开发前影响分析", "### B. 开发后实现增强", "旧系统基线", "不得把旧代码当成新需求实现")

    def test_07_postdevelopment_diff_updates_cases_and_regression(self):
        self.assert_terms_in("### B. 开发后实现增强", "### C. 阶段无法确认", "新增、修改、删除分支", "回归范围")

    def test_08_requirement_wins_when_code_conflicts(self):
        self.assert_terms_in("## 最高优先级铁律", "## 定位与边界", "需求决定系统应该怎样，代码说明系统当前怎样实现", "需求与代码冲突", "实现差异")

    def test_09_missing_implementation_layer_is_an_evidence_gap(self):
        self.assert_terms_in("## 输入准入与互斥路由", "## 双重输入分级", "标记证据缺口", "不编造缺失层", "不停止其他分析")

    def test_10_unrelated_code_is_not_deep_read(self):
        self.assert_terms_in("## 固定工作流", "## 需求相关全链路定位", "全仓搜索", "相关性筛选", "相关链路深读", "停止无关扩展")

    def test_11_history_is_relevant_deduplicated_regression_only(self):
        self.assert_terms_in("## 历史用例、历史问题与待动态验证", "## 正式测试用例格式", "历史回归", "同根因不得重复生成", "无关历史问题直接排除")

    def test_12_static_suspicion_is_dynamic_verification_not_bug(self):
        self.assert_terms_in("## 历史用例、历史问题与待动态验证", "## 正式测试用例格式", "待动态验证", "不得直接标记为Bug")
        self.assert_terms_in("## 最高优先级铁律", "## 定位与边界", "未经执行不得认定为Bug")

    def test_13_formal_case_contract_is_exactly_ten_columns(self):
        contract = (ROOT / "skill-sources" / "shared" / "requirement-test-case-output-contract.md").read_text(encoding="utf-8")
        for column in CASE_COLUMNS:
            self.assertIn(column, contract)
        self.assertIn("用例 ID 使用阿拉伯数字，从 1 开始连续", contract)
        self.assertIn("执行结果默认填写“未执行”", contract)
        self.assertNotIn("| 实际结果 |", contract)

    def test_14_xlsx_html_and_simhei_smoke_contract(self):
        self.assert_terms("正式测试用例", "Excel", "HTML", "SimHei", "可编辑")

    def test_15_generated_package_is_self_contained_and_installable(self):
        build_all(ROOT)
        manifest = {item["slug"]: item for item in load_manifest(ROOT)["skills"]}
        self.assertIn(SLUG, manifest)
        self.assertTrue((PACKAGE / "SKILL.md").is_file())
        self.assertTrue((PACKAGE / "agents" / "openai.yaml").is_file())
        self.assertTrue((PACKAGE / "scripts" / "render-test-assets.mjs").is_file())
        self.assertTrue((PACKAGE / "scripts" / "render-test-assets.py").is_file())
        schema = PACKAGE / "assets" / "graybox-report.schema.json"
        template = PACKAGE / "assets" / "graybox-report-template.json"
        self.assertTrue(schema.is_file())
        self.assertTrue(template.is_file())
        schema_data = json.loads(schema.read_text(encoding="utf-8"))
        template_data = json.loads(template.read_text(encoding="utf-8"))
        self.assertEqual("SimHei", schema_data["properties"]["excel_font"]["const"])
        self.assertEqual("SimHei", template_data["excel_font"])
        self.assertEqual(CASE_COLUMNS, template_data["sheets"][0]["columns"])
        generated_meta, generated_body = parse_frontmatter((PACKAGE / "SKILL.md").read_text(encoding="utf-8"))
        self.assertEqual(SLUG, generated_meta["name"])
        self.assertNotIn("../requirement-test-workbench", generated_body)
        self.assertNotIn("skills/requirement-test-workbench", generated_body)

        with tempfile.TemporaryDirectory() as directory:
            install_root = Path(directory) / "installed"
            result = subprocess.run(
                [
                    "powershell.exe",
                    "-NoProfile",
                    "-ExecutionPolicy",
                    "Bypass",
                    "-File",
                    str(ROOT / "scripts" / "install.ps1"),
                    "-Skill",
                    SLUG,
                    "-SourceDirectory",
                    str(ROOT),
                    "-InstallRoot",
                    str(install_root),
                ],
                capture_output=True,
                check=False,
            )
            self.assertEqual(0, result.returncode, (result.stdout + result.stderr).decode(errors="replace"))
            installed = install_root / SLUG
            self.assertEqual((PACKAGE / "SKILL.md").read_bytes(), (installed / "SKILL.md").read_bytes())
            self.assertTrue((installed / "scripts" / "render-test-assets.mjs").is_file())
            self.assertTrue((installed / "scripts" / "render-test-assets.py").is_file())
            self.assertTrue((installed / "assets" / "graybox-report.schema.json").is_file())
            self.assertTrue((installed / "assets" / "graybox-report-template.json").is_file())

    def test_17_contract_has_no_known_gate_or_truth_contradictions(self):
        forbidden = [
            "L0 可以生成正式测试用例",
            "代码决定系统应该怎样",
            "代码完整即可绕过 P0",
            "代码疑点直接标记为Bug",
            "只有需求文档时扫描代码仓库",
        ]
        for phrase in forbidden:
            self.assertNotIn(phrase, self.text)

    def test_16_workbench_and_graybox_share_the_source_contract(self):
        contract = ROOT / "skill-sources" / "shared" / "requirement-test-case-output-contract.md"
        self.assertTrue(contract.is_file())
        marker = "<!-- include:requirement-test-case-output-contract -->"
        self.assertIn(marker, SOURCE.read_text(encoding="utf-8"))
        self.assertIn(marker, (ROOT / "根据需求-用例生成_skill.md").read_text(encoding="utf-8"))


class EnhancedGrayboxRendererSmokeTest(unittest.TestCase):
    def test_exact_columns_openable_xlsx_and_browsable_html(self):
        report = {
            "title": "增强版灰盒测试用例",
            "generated_at": "2026-07-22T00:00:00Z",
            "excel_font": "SimHei",
            "skill_invocation": {"primary": SLUG},
            "sheets": [
                {
                    "name": "正式测试用例",
                    "kind": "test_cases",
                    "columns": CASE_COLUMNS,
                    "rows": [
                        {
                            "values": [
                                "1",
                                "权限",
                                "服务端校验数据归属",
                                "1. 验证绕过前端后仍校验对象归属",
                                "1. 已准备两个不同归属账号",
                                "1. 账号A请求账号B对象",
                                "1. 请求被拒绝且数据不变化",
                                "P0",
                                "未执行",
                                "实现证据：Controller#update",
                            ]
                        }
                    ],
                }
            ],
        }
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "report.json"
            source.write_text(json.dumps(report, ensure_ascii=False), encoding="utf-8")
            result = subprocess.run(
                [
                    "node",
                    str(ROOT / "tooling" / "test-case-renderer.mjs"),
                    "--input",
                    str(source),
                    "--output-dir",
                    str(root),
                    "--basename",
                    "graybox",
                ],
                capture_output=True,
                text=True,
                encoding="utf-8",
                check=False,
                env={**os.environ, "TESTING_SKILLS_FORCE_PORTABLE_XLSX": "1"},
            )
            self.assertEqual(0, result.returncode, result.stderr)
            xlsx = root / "graybox.xlsx"
            html = root / "graybox.html"
            self.assertTrue(xlsx.is_file())
            self.assertTrue(html.is_file())
            with zipfile.ZipFile(xlsx) as archive:
                self.assertIn("xl/workbook.xml", archive.namelist())
                styles = archive.read("xl/styles.xml").decode("utf-8")
                sheet = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")
            self.assertIn('name val="SimHei"', styles)
            self.assertIn("执行结果（通过 / 不通过 / 未执行）", sheet)
            html_text = html.read_text(encoding="utf-8")
            self.assertIn("正式测试用例", html_text)
            self.assertIn("localStorage", html_text)


class EnhancedGrayboxOrderExportExcelRegressionTest(unittest.TestCase):
    def setUp(self):
        self.report = order_export_report()

    def _write_report(self, root):
        source = root / "后台订单列表导出.json"
        source.write_text(json.dumps(self.report, ensure_ascii=False), encoding="utf-8")
        return source

    def _assert_order_export_regression(self, xlsx, html):
        assert_workbook_matches_report(self, xlsx, self.report)
        assert_html_matches_report(self, html, self.report)
        workbook = load_workbook(xlsx, data_only=False)
        sheet = workbook["正式测试用例"]
        self.assertEqual("A1:J38", sheet.calculate_dimension())
        self.assertEqual("用例 ID", sheet["A1"].value)
        self.assertEqual("所属模块", sheet["B1"].value)
        self.assertEqual("备注", sheet["J1"].value)
        self.assertEqual("【模块分割行】", sheet["A2"].value)
        self.assertEqual("1", sheet["A3"].value)
        self.assertEqual("运营人员按当前列表条件发起导出", sheet["C3"].value)
        ids = [
            str(sheet.cell(row=row, column=1).value)
            for row in range(2, sheet.max_row + 1)
            if sheet.cell(row=row, column=1).value != "【模块分割行】"
        ]
        self.assertEqual([str(value) for value in range(1, 31)], ids)
        statuses = [
            sheet.cell(row=row, column=9).value
            for row in range(2, sheet.max_row + 1)
            if sheet.cell(row=row, column=1).value != "【模块分割行】"
        ]
        self.assertEqual(["未执行"] * 30, statuses)
        self.assertNotEqual((380, 1), (sheet.max_row, sheet.max_column))
        self.assertTrue(all(
            value is None or not isinstance(value, str) or len(value) != 1 or value in {"1", "2", "3", "4", "5", "6", "7", "8", "9"}
            for row in sheet.iter_rows()
            for cell in row
            for value in [cell.value]
            if cell.coordinate in {"A1", "A2", "C3"}
        ))

    def test_node_renderer_preserves_order_export_sheet_matrices(self):
        node = shutil.which("node")
        self.assertIsNotNone(node, "Node renderer regression requires Node.js")
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = self._write_report(root)
            result = subprocess.run(
                [node, str(ROOT / "tooling" / "test-case-renderer.mjs"), "--input", str(source), "--output-dir", str(root), "--basename", "order-export"],
                capture_output=True,
                text=True,
                encoding="utf-8",
                check=False,
                env={**os.environ, "TESTING_SKILLS_FORCE_PORTABLE_XLSX": "1"},
            )
            self.assertEqual(0, result.returncode, result.stderr)
            verifier = subprocess.run(
                [
                    sys.executable,
                    str(ROOT / "skill-sources" / SLUG / "scripts" / "render-test-assets.py"),
                    "--verify-only",
                    "--input",
                    str(source),
                    "--xlsx",
                    str(root / "order-export.xlsx"),
                    "--html",
                    str(root / "order-export.html"),
                ],
                capture_output=True,
                text=True,
                encoding="utf-8",
                check=False,
                env={**os.environ, "PYTHONUTF8": "1", "PYTHONIOENCODING": "utf-8"},
            )
            self.assertEqual(0, verifier.returncode, verifier.stderr)
            self.assertTrue(json.loads(verifier.stdout)["verified"])
            self._assert_order_export_regression(root / "order-export.xlsx", root / "order-export.html")

    def test_python_fallback_preserves_order_export_sheet_matrices_without_node(self):
        fallback = ROOT / "skill-sources" / SLUG / "scripts" / "render-test-assets.py"
        self.assertTrue(fallback.is_file(), "Node 缺失时必须使用包内受控 Python 备用生成器")
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = self._write_report(root)
            result = subprocess.run(
                [sys.executable, str(fallback), "--input", str(source), "--output-dir", str(root), "--basename", "order-export"],
                capture_output=True,
                text=True,
                encoding="utf-8",
                check=False,
                env={**os.environ, "PATH": ""},
            )
            self.assertEqual(0, result.returncode, result.stderr)
            self._assert_order_export_regression(root / "order-export.xlsx", root / "order-export.html")

    def test_fallback_dependency_failure_keeps_json_and_html_without_writing_xlsx(self):
        fallback = ROOT / "skill-sources" / SLUG / "scripts" / "render-test-assets.py"
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = self._write_report(root)
            html = root / "order-export.html"
            html.write_text("保留已有 HTML", encoding="utf-8")
            result = subprocess.run(
                [sys.executable, "-S", str(fallback), "--input", str(source), "--output-dir", str(root), "--basename", "order-export"],
                capture_output=True,
                text=True,
                encoding="utf-8",
                check=False,
                env={**os.environ, "PYTHONUTF8": "1", "PYTHONIOENCODING": "utf-8"},
            )
            self.assertNotEqual(0, result.returncode)
            self.assertIn("未生成或宣称校验通过的 Excel", result.stderr)
            self.assertFalse((root / "order-export.xlsx").exists())
            self.assertEqual("保留已有 HTML", html.read_text(encoding="utf-8"))

    def test_skill_calibrates_priority_and_unknown_order_export_expectations(self):
        text = SOURCE.read_text(encoding="utf-8")
        self.assertIn("P0 不得因属于本轮核心功能而批量泛化", text)
        self.assertIn("列表与导出时间边界一致", text)
        self.assertIn("具体开闭区间待确认", text)
        self.assertIn("无数据时只保留一个可判定预期", text)
        self.assertIn("不得并列两个互斥结果", text)


if __name__ == "__main__":
    unittest.main()
